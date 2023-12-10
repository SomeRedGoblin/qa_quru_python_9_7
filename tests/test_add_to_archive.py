import csv
import os
import shutil
from zipfile import ZipFile, ZIP_STORED

from PyPDF2 import PdfReader
from openpyxl import load_workbook

from conftest import TMP_DIR, TEST_DATA_DIR, CURRENT_DIR

files_dir = os.listdir(TEST_DATA_DIR)


# Данный способ упаковывает вместе с путем до файла. Не придумал как использовать arcname
def test_add_files_to_zip_via_shutil():
    shutil._make_zipfile(base_name=f'{TMP_DIR}\hello', base_dir=TEST_DATA_DIR, )
    assert os.path.isfile(f'{TMP_DIR}\hello.zip')


def test_add_files_to_zip_via_zipfile():
    if os.path.isfile(f'{TMP_DIR}\hello.zip'):
        shutil.rmtree(TMP_DIR)

    os.mkdir(f'{CURRENT_DIR}\\resources')

    with ZipFile(f'{TMP_DIR}\hello.zip', mode='w', compression=ZIP_STORED) as zip_file:
        for file in files_dir:
            add_file = os.path.join(TEST_DATA_DIR, file)
            zip_file.write(add_file, arcname=f'{file}')
    assert os.path.isfile(f'{TMP_DIR}\hello.zip')
    assert zip_file.namelist() == ['file_example_CSV_500.csv', 'file_example_XLSX_10.xlsx', 'pdf-test.pdf']


def test_read_csv_file():
    # Проверка не распаковывая
    with ZipFile(f'{TMP_DIR}\hello.zip') as zip_file:
        print(zip_file.namelist())
        text = zip_file.read('file_example_CSV_500.csv')

        assert text.__len__() == 27950
        assert text.startswith(str.encode(',First Name,Last Name,Gender,Country,Age,Date,Id'))

    # Проверка после распаковки
    with ZipFile(f'{TMP_DIR}\hello.zip') as zip_file:
        zip_file.extract('file_example_CSV_500.csv', path="tmp")

    with open('tmp/file_example_CSV_500.csv') as f:
        reader = csv.reader(f)
        assert ['410', 'Fallon', 'Winward', 'Female', 'Great Britain', '28', '16/08/2016', '5486'] in reader


def test_read_pdf_file():
    # Проверка не распаковывая
    with ZipFile(f'{TMP_DIR}\hello.zip') as zip_file:
        print(zip_file.namelist())
        text = zip_file.read('pdf-test.pdf')

        assert text.__len__() == 20597

    # Проверка после распаковки
    with ZipFile(f'{TMP_DIR}\hello.zip') as zip_file:
        zip_file.extract('pdf-test.pdf', path="tmp")

    reader = PdfReader('tmp/pdf-test.pdf')
    assert len(reader.pages) == 1

    page = reader.pages[0]
    text = page.extract_text()
    print(text)
    assert text.__contains__('Congratulations, your comput er is equipped with a PDF (Portable Document Format)')


def test_read_xlsx_file():
    # Проверка не распаковывая
    with ZipFile(f'{TMP_DIR}\hello.zip') as zip_file:
        print(zip_file.namelist())
        text = zip_file.read('file_example_XLSX_10.xlsx')

        assert text.__len__() == 5425

    # Провверка после распаковки
    with ZipFile(f'{TMP_DIR}\hello.zip') as zip_file:
        zip_file.extract('file_example_XLSX_10.xlsx', path="tmp")

    # открываем файл
    workbook = load_workbook('tmp/file_example_XLSX_10.xlsx')
    sheet = workbook.active
    assert sheet.max_column == 8
    assert sheet.max_row == 10
    assert sheet.cell(row=3, column=2).value == "Mara"
